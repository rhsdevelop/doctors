import openpyxl
import smtplib
from unidecode import unidecode

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.template import loader

from .forms import (EmailConfigForm, AddDoctorForm, AddPhoneForm, AddSpecialtyForm, AddVisitForm, FindDoctorForm,
                    FindSpecialtyForm, FindVisitForm, PlanilhaEmergenciaForm, FindPlanilhaForm, FilterGvpStatusForm, GvpVisitForm)
from .models import EmailConfiguration, Doctor, Phone, Specialty, Visit, PlanilhaEmergencia, GvpVisit
from .utils import disparar_alerta_gvp

# Create your views here.


@login_required
@permission_required('admin.can_change_email_config', raise_exception=True)
def configure_email(request):
    # Tenta pegar a primeira configuração existente ou cria uma nova
    config = EmailConfiguration.objects.first()
    
    if request.method == 'POST':
        form = EmailConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configurações de e-mail atualizadas!")
            return redirect('/')
    else:
        form = EmailConfigForm(instance=config)
    
    return render(request, 'config/email_setup.html', {
        'form': form,
        'title': 'Configurar Servidor de E-mail'
    })

@login_required
@permission_required('admin.can_change_email_config', raise_exception=True)
def testar_smtp(request):
    # Pega os dados enviados pelo formulário via GET ou POST (AJAX)
    host = request.GET.get('smtp_server')
    port = request.GET.get('smtp_port')
    user = request.GET.get('email_user')
    password = request.GET.get('email_password')
    use_tls = request.GET.get('use_tls') == 'true'

    try:
        # Tenta a conexão SMTP
        server = smtplib.SMTP(host, int(port), timeout=10)
        if use_tls:
            server.starttls()
        
        server.login(user, password)
        server.quit()
        
        return JsonResponse({'status': 'success', 'message': 'Conexão e Autenticação bem-sucedidas!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Falha na conexão: {str(e)}'})

@login_required
def index(request):
    doctors_list = Doctor.objects.order_by('id')
    template = loader.get_template('index.html')
    context = {
        'title': 'Gestão de Médicos Cooperadores da Colih',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'doctors_list': doctors_list,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.add_specialty', raise_exception=True)
def add_specialty(request):
    if request.POST:
        form = AddSpecialtyForm(request.POST)
        item = form
        item.save()
        messages.success(request, 'Registro adicionado com sucesso.')
        return redirect('/specialties/list')
    form = AddSpecialtyForm()
    template = loader.get_template('specialties/add.html')
    context = {
        'title': 'Adicionar Especialidade Médica',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.change_specialty', raise_exception=True)
def edit_specialty(request, specialty_id):
    specialty = Specialty.objects.get(id=specialty_id)
    if request.POST:
        form = AddSpecialtyForm(request.POST, instance=specialty)
        item = form
        item.save()
        messages.success(request, 'Registro alterado com sucesso.')
        return redirect('/specialties/list')
    form = AddSpecialtyForm(instance=specialty)
    template = loader.get_template('specialties/edit.html')
    context = {
        'title': 'Editar Especialidade Cadastrada',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.view_specialty', raise_exception=True)
def list_specialties(request):
    form = FindSpecialtyForm(request.GET)
    form.fields['name'].required = False
    filter_search = {}
    for key, value in request.GET.items():
        if key in ['name', 'hospital', 'city'] and value:
            filter_search['%s__icontains' % key] = value
    specialties = Specialty.objects.filter(**filter_search)
    template = loader.get_template('specialties/list.html')
    context = {
        'title': 'Relação de Especialidades Médicas',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'specialties': specialties,
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.add_doctor', raise_exception=True)
def add_doctor(request):
    if request.POST:
        form = AddDoctorForm(request.POST)
        item = form
        item.save()
        messages.success(request, 'Registro adicionado com sucesso.')
        return redirect('/doctors/list')
    form = AddDoctorForm()
    template = loader.get_template('doctors/add.html')
    context = {
        'title': 'Adicionar Médico Cooperador',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.change_doctor', raise_exception=True)
def edit_doctor(request, doctor_id):
    doctor = Doctor.objects.get(id=doctor_id)
    if request.POST:
        if 'number' in request.POST:
            form = AddPhoneForm(request.POST)
            item = form.save(commit=False)
            if not item.number:
                messages.error(request, 'Informe o número de telefone a ser adicionado!')
            else:
                item.doctor = doctor
                item.save()
                messages.success(request, 'Contato telefônico adicionado com sucesso.')
        else:
            form = AddDoctorForm(request.POST, instance=doctor)
            item = form
            item.save()
            messages.success(request, 'Registro alterado com sucesso.')
            return redirect('/doctors/list')
    phones = Phone.objects.filter(doctor=doctor)
    form = AddDoctorForm(instance=doctor)
    phone_form = AddPhoneForm()
    template = loader.get_template('doctors/edit.html')
    context = {
        'title': 'Dados de Médico Cadastrado',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'phones': phones,
        'form': form,
        'doctor_id': doctor_id,
        'phone_form': phone_form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.view_doctor', raise_exception=True)
def export_doctors_xlsx(request):
    filter_search = {}
    for key, value in request.GET.items():
        if key in ['name', 'hospital', 'city'] and value:
            filter_search['%s__icontains' % key] = value
        elif key in ['specialty'] and value:
            filter_search[key] = value
    doctors = Doctor.objects.filter(**filter_search).select_related('specialty', 'hospital')


    # 1. Criar o Workbook e a planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Médicos"

    # 2. Definir o cabeçalho
    columns = [
        'Nome completo do Médico', 
        'Especialidade 1', 
        'Especialidade 2 (opcional)', 
        'Especialidade 3 (opcional)', 
        'Subespecialidade (opcional)', 
        'Tipo de paciente', 
        'Atende SUS?', 
        'Faz cirurgias?', 
        'Data da última visita ', 
        'É Testemunha de Jeová?', 
        'É Consultor  informado ao HID?', 
    ]
    ws.append(columns)

    # Estilizar o cabeçalho (Negrito e Cor de fundo)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")

    # 3. Buscar os dados do banco
    # Usamos select_related para performance nas chaves estrangeiras
    doctors = Doctor.objects.filter(**filter_search).select_related('specialty', 'hospital')

    for doctor in doctors:
        ws.append([
            doctor.name,
            doctor.specialty.name if doctor.specialty else "",
            doctor.specialty2.name if doctor.specialty2 else "",
            doctor.specialty3.name if doctor.specialty3 else "",
            doctor.subspecialty if doctor.subspecialty else "",
            doctor.type_patient,
            'Sim'if doctor.attends_sus else 'Não',
            'Sim'if doctor.performs_surgeries else 'Não',
            'Sim'if doctor.last_visit else 'Não',
            'Sim'if doctor.is_jehovah_witness else 'Não',
            'Sim'if doctor.is_hid_consultant else 'Não',
        ])

    # 4. Ajustar a largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    # 5. Preparar a resposta do navegador
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Medicos.xlsx"'
    
    wb.save(response)
    return response

@login_required
@permission_required('doctors.view_doctor', raise_exception=True)
def list_doctors(request):
    form = FindDoctorForm(request.GET)
    form.fields['name'].required = False
    form.fields['specialty'].required = False
    form.fields['hospital'].required = False
    form.fields['city'].required = False
    filter_search = {}
    for key, value in request.GET.items():
        if key in ['name', 'hospital', 'city'] and value:
            filter_search['%s__icontains' % key] = value
        elif key in ['specialty'] and value:
            filter_search[key] = value
    doctors = Doctor.objects.filter(**filter_search)
    template = loader.get_template('doctors/list.html')
    context = {
        'title': 'Relação de Médicos Cooperadores',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'doctors': doctors,
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.delete_phone', raise_exception=True)
def delete_phone(request, doctor_id, phone_id):
    phone = Phone.objects.get(id=phone_id)
    phone.delete()
    messages.success(request, 'Contato telefônico removido com sucesso.')
    return redirect('/doctors/%s/edit/' % doctor_id)

@login_required
@permission_required('doctors.view_phone', raise_exception=True)
def list_phones(request):
    phones = Phone.objects.all()
    print(phones)
    template = loader.get_template('phones/list.html')
    context = {
        'title': 'Contatos telefônicos dos médicos',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'phones': phones,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.view_doctor', raise_exception=True)
def detail(request, doctor_id):
    try:
        doctor = Doctor.objects.get(pk=doctor_id)
    except Doctor.DoesNotExist:
        raise Http404("Doctor does not exist")
    context = {
        'title': 'Dados de Médico Cadastrado',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'doctor': '%s - %s' % (doctor.name, doctor.hospital)
    }
    return render(request, 'doctors/detail.html', context)

@login_required
@permission_required('doctors.view_visit', raise_exception=True)
def list_visits(request):
    # 1. Instancia o formulário com os dados da busca (GET)
    form = FindVisitForm(request.GET)
    
    # 2. Remove a obrigatoriedade dos campos para o filtro funcionar vazio
    for field in form.fields:
        form.fields[field].required = False

    # 3. Monta o dicionário de filtros
    filter_search = {}
    
    if form.is_valid():
        # Pegamos os dados limpos do formulário
        data = form.cleaned_data
        
        # Filtro por nome do médico (relacionamento doctor -> name)
        if data.get('doctor_name'):
            filter_search['doctor__name__icontains'] = data['doctor_name']
        
        # Filtro por tipo de visita (exato)
        if data.get('visit_type'):
            filter_search['visit_type'] = data['visit_type']
            
        # Filtro por especialidade (exato ou nome)
        if data.get('specialty'):
            filter_search['specialty'] = data['specialty']

    # 4. Busca no banco de dados com os filtros aplicados
    # select_related otimiza a consulta trazendo os dados do médico e especialidade juntos
    visits = Visit.objects.filter(**filter_search).select_related('doctor', 'specialty').order_by('-visit_date')

    template = loader.get_template('visits/list.html')
    context = {
        'title': 'Histórico de Visitas',
        'username': '%s %s' % (request.user.first_name, request.user.last_name),
        'visits': visits,
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.add_visit', raise_exception=True)
def add_visit(request):
    if request.method == 'POST':
        # Se o usuário enviou dados (clicou em Salvar)
        form = AddVisitForm(request.POST)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Visita registrada com sucesso!')
            # Redireciona para a mesma página para cadastrar outra, 
            # ou muda para a lista de visitas (ex: return redirect('list_visits'))
            return redirect('/visits/list') 
        else:
            messages.error(request, 'Erro ao registrar visita. Verifique os campos.')
    else:
        # Se for apenas um acesso à página (GET), cria o formulário vazio
        form = AddVisitForm()

    context = {
        'form': form
    }
    
    # Certifique-se de criar este arquivo HTML no passo seguinte
    return render(request, 'visits/add.html', context)

@login_required
@permission_required('doctors.change_visit', raise_exception=True)
def edit_visit(request, visit_id):
    # 1. Busca a visita pelo ID. Se não achar, dá erro 404 automaticamente.
    visit = get_object_or_404(Visit, id=visit_id)

    if request.method == 'POST':
        # 2. Carrega o formulário com os dados enviados (POST) E a instância atual
        form = AddVisitForm(request.POST, instance=visit)
        
        if form.is_valid():
            form.save() # Como tem 'instance', ele faz um UPDATE, não um INSERT
            messages.success(request, 'Visita alterada com sucesso.')
            return redirect('/visits/list') # Ou return redirect('/visits/list')
        else:
            messages.error(request, 'Erro ao alterar visita. Verifique os campos.')
            
    else:
        # 3. Se for GET, carrega o formulário preenchido com os dados do banco
        form = AddVisitForm(instance=visit)

    context = {
        'title': 'Editar Visita',
        'form': form,
        # Se precisares do username no template como no exemplo:
        'username': f"{request.user.first_name} {request.user.last_name}" 
    }
    
    # Podemos reutilizar o template de adicionar ou criar um específico 'edit.html'
    return render(request, 'visits/add.html', context)

@login_required
@permission_required('doctors.delete_visit', raise_exception=True)
def delete_visit(request, visit_id):
    # 1. Busca a visita ou dá erro 404
    visit = get_object_or_404(Visit, id=visit_id)

    if request.method == 'POST':
        # 2. Se o formulário foi enviado (confirmou a exclusão)
        visit.delete()
        messages.success(request, 'Visita excluída com sucesso.')
        return redirect('/visits/list')

    # 3. Se for GET, mostra a página de confirmação
    context = {
        'title': 'Excluir Visita',
        'visit': visit
    }
    return render(request, 'visits/delete.html', context)

@login_required
@permission_required('doctors.add_planilhaemergencia', raise_exception=True)
def add_planilha_emergencia(request):
    if request.method == 'POST':
        form = PlanilhaEmergenciaForm(request.POST)
        if form.is_valid():
            planilha = form.save()
            messages.success(request, f'Planilha do paciente {planilha.nome_paciente} salva com sucesso!')
            # Você pode redirecionar para a lista de planilhas ou para o detalhe desta planilha
            return redirect('/emergencia/list') 
        else:
            messages.error(request, 'Erro ao salvar formulário. Verifique os campos em vermelho.')
    else:
        form = PlanilhaEmergenciaForm()

    context = {
        'form': form,
        'title': 'Nova Planilha de Emergência'
    }
    
    # O template precisa ser criado (veja instrução abaixo)
    return render(request, 'emergency/add.html', context)

@login_required
@permission_required('doctors.change_planilhaemergencia', raise_exception=True)
def edit_planilha_emergencia(request, planilha_id):
    # 1. Busca a planilha existente pelo ID
    planilha = get_object_or_404(PlanilhaEmergencia, id=planilha_id)

    if request.method == 'POST':
        # 2. Passamos o POST e a instância (registro atual) para o form
        form = PlanilhaEmergenciaForm(request.POST, instance=planilha)
        
        if form.is_valid():
            form.save()
            messages.success(request, f'Planilha do paciente {planilha.nome_paciente} atualizada com sucesso!')
            return redirect('/emergencia/list/') # Ajuste para sua URL de listagem
        else:
            messages.error(request, 'Erro ao atualizar a planilha. Verifique os campos abaixo.')
    else:
        # 3. No GET, o form vem preenchido com os dados da 'planilha'
        form = PlanilhaEmergenciaForm(instance=planilha)

    context = {
        'form': form,
        'title': f'Editar Planilha: {planilha.nome_paciente}',
        'username': f"{request.user.first_name} {request.user.last_name}"
    }
    
    # Reutilizamos o mesmo template de adição para manter o layout de abas
    return render(request, 'emergency/add.html', context)

@login_required
@permission_required('doctors.view_planilhaemergencia', raise_exception=True)
def list_planilhas(request):
    form = FindPlanilhaForm(request.GET)
    filter_search = {}

    if form.is_valid():
        data = form.cleaned_data
        
        # Filtro por texto (contém)
        if data.get('nome_paciente'):
            filter_search['nome_paciente__icontains'] = data['nome_paciente']
        
        # Filtro exato (ID do hospital)
        if data.get('nome_hospital'):
            filter_search['nome_hospital'] = data['nome_hospital']

    # Busca no banco ordenando da mais recente para a mais antiga
    planilhas = PlanilhaEmergencia.objects.filter(**filter_search)\
        .select_related('nome_hospital')\
        .order_by('-data_hora_contato')

    template = loader.get_template('emergency/list.html')
    context = {
        'title': 'Gestão de Planilhas de Emergência',
        'planilhas': planilhas,
        'form': form,
    }
    return HttpResponse(template.render(context, request))

@login_required
@permission_required('doctors.change_planilhaemergencia', raise_exception=True)
def submeter_para_gvp(request, planilha_id):
    # Busca a planilha
    planilha = get_object_or_404(PlanilhaEmergencia, id=planilha_id)
    
    # Altera o status para 'Em Acompanhamento'
    planilha.status_gvp = 'AND'
    planilha.save()

    # Dispara o e-mail automático
    enviado = disparar_alerta_gvp(planilha, request)

    if enviado:
        messages.success(request, f'Sucesso! {planilha.nome_paciente} enviado ao GVP e equipe notificada por e-mail.')
    else:
        messages.warning(request, f'Status alterado, mas houve uma falha ao enviar o e-mail de alerta.')
    
    # Redireciona de volta para a lista onde ele estava
    return redirect('/emergencia/list/')

@login_required
@permission_required('doctors.change_planilhaemergencia', raise_exception=True)
def gerar_boletim_whatsapp(request, planilha_id):
    p = get_object_or_404(PlanilhaEmergencia, id=planilha_id)
    
    # Montagem do texto
    texto = (
        f"*📋 BOLETIM DE ACOMPANHAMENTO - GVP*\n\n"
        f"*👤 PACIENTE:* {p.nome_paciente}\n"
        f"*🏥 HOSPITAL:* {p.nome_hospital.name if p.nome_hospital else 'Não informado'}\n"
        f"*🚪 QUARTO:* {p.numero_quarto or '-'}\n"
        f"*🩺 MÉDICO:* {p.medico_responsavel}\n"
        f"*📝 DIAGNÓSTICO:* {p.problema_especifico[:200]}...\n\n"
        f"*💡 ESTRATÉGIA:* {p.estrategias_opcoes[:200]}...\n\n"
        f"*📍 SITUAÇÃO GVP:* {p.get_status_gvp_display()}\n"
        f"------------------------------------------\n"
        f"_Gerado via Sistema COLIH/GVP em {p.updated_at.strftime('%d/%m/%Y %H:%M')} por {request.user.first_name + ' ' + request.user.last_name}_"
    )
    
    return JsonResponse({'texto': texto})

@login_required
@permission_required('doctors.view_gvpvisit', raise_exception=True)
def list_gvp_active_cases(request):
    form = FilterGvpStatusForm(request.GET)
    # Filtro base: exclui os 'PEN' (Pendentes)
    filter_search = {'status_gvp__in': ['AND', 'FIN']}
    title = 'Acompanhamentos GVP (Ativos e Concluídos)'
    if not request.user.is_superuser and request.user.groups.filter(name='GVP - Operacional').exists():
        my_active = GvpVisit.objects.filter(designated_members=request.user).values_list('planilha__id', flat=True)
        filter_search['id__in'] = my_active
        title = 'Meus Acompanhamentos GVP (Ativos e Concluídos)'
    if form.is_valid():
        if form.cleaned_data.get('nome_paciente'):
            filter_search['nome_paciente__icontains'] = form.cleaned_data['nome_paciente']
        if form.cleaned_data.get('status_gvp'):
            filter_search['status_gvp'] = form.cleaned_data['status_gvp']
        if form.cleaned_data.get('hospital'):
            filter_search['nome_hospital'] = form.cleaned_data['hospital']

    planilhas = PlanilhaEmergencia.objects.filter(**filter_search).select_related('nome_hospital')
    context = {
        'title': title,
        'planilhas': planilhas,
        'form': form,
    }
    return render(request, 'gvp/list_plan.html', context)

@login_required
@permission_required('doctors.add_gvpvisit', raise_exception=True)
def add_gvp_visit(request, planilha_id=None):
    # Se passarmos o ID pela URL, já buscamos a planilha para exibir os dados ao GVP
    planilha = None
    if planilha_id:
        planilha = get_object_or_404(PlanilhaEmergencia, id=planilha_id)

    if request.method == 'POST':
        # Busca se já existe uma visita para esta planilha para atualizar em vez de duplicar
        gvpvisit = GvpVisit.objects.filter(planilha=planilha).first()
        form = GvpVisitForm(request.POST, instance=gvpvisit, user=request.user)
        if form.is_valid():
            viva_gvp = form.save()
            # Lógica de finalização de status
            if form.cleaned_data.get('finalizar_caso'):
                planilha.status_gvp = 'FIN'
                messages.info(request, "Caso Finalizado.")
            else:
                # Se desmarcar, ele volta para 'Em Andamento' (opcional)
                planilha.status_gvp = 'AND'
            planilha.save()
            messages.success(request, f'Ação do GVP registrada para o paciente {viva_gvp.planilha.nome_paciente}.')
            return redirect('/gvp/acompanhamentos/') # Ou para uma lista própria do GVP
        else:
            # Carregamento inicial do formulário
            gvpvisit = GvpVisit.objects.filter(planilha=planilha).first()
            form = GvpVisitForm(instance=gvpvisit)
            if planilha:
                # Trava o queryset da planilha e remove o empty_label
                form.fields['planilha'].queryset = PlanilhaEmergencia.objects.filter(id=planilha.id)
                form.fields['planilha'].empty_label = None
                form.fields['planilha'].initial = planilha.id
    else:
        # Se viemos de uma planilha específica, já deixamos ela selecionada no form
        if planilha:
            gvpvisit = GvpVisit.objects.filter(planilha=planilha).first()
            if gvpvisit:
                form = GvpVisitForm(instance=gvpvisit, user=request.user)                
            else:
                initial_data = {'planilha': planilha.id} if planilha else {}
                print(initial_data)
                form = GvpVisitForm(initial=initial_data, user=request.user)
            form.fields['planilha'].queryset = PlanilhaEmergencia.objects.filter(id=planilha.id)
            # 2. Remove a opção "---------" (Label vazio)
            form.fields['planilha'].empty_label = None
            # 3. Garante que o campo esteja selecionado (reforço)
            form.fields['planilha'].initial = planilha.id
        else:
            form = GvpVisitForm()

    context = {
        'form': form,
        'planilha': planilha, # Passamos a planilha para mostrar os dados de leitura
        'title': 'Registro de Atuação GVP'
    }
    return render(request, 'gvp/add.html', context)